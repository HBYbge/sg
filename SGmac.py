#TODO 导入
import tkinter as tk,json,hashlib,os,time,shutil,webbrowser as web,openpyxl
from tkinter import ttk,messagebox as mb,filedialog as fd,simpledialog as sd
from openpyxl import utils,styles
import platform

if platform.system() == 'Darwin':
    import sys
    try:
        from tkinter import font
        default_font = font.nametofont("TkDefaultFont")
        default_font.configure(family="Helvetica", size=11)
    except:
        pass
    if sys.version_info[0] == 3 and sys.version_info[1] >= 8:
        try:
            import tkinter
            tkinter._default_root = None
        except:
            pass

if not os.path.exists('info'):
    os.makedirs('info')

required_files = ['user.json', 'student.json', 'score.json', 'grade.json', 'class.json', 'subject.json', 'exam.json']
for file in required_files:
    filepath = os.path.join('info', file)
    if not os.path.exists(filepath):
        with open(filepath, 'w', encoding='utf-8') as f:
            if file == 'user.json':
                json.dump({}, f)
            else:
                f.write('{}')

#TODO 函数
def sign_up():
    def create_account():
        nonlocal e
        if p.get() == p2.get():
            try:
                hash_p = hashlib.sha256(p.get().encode('utf-8')).hexdigest()
                json_file = "info/user.json"
                d = {}

                if os.path.exists(json_file):
                    with open(json_file, "r", encoding="utf-8") as f:
                        file_content = f.read().strip()
                        if file_content:
                            d = json.loads(file_content)

                data = {**d, n.get(): {'password':hash_p,'email':e.get()}}

                with open(json_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)

                print(hash_p)

            except Exception as e:
                with open(json_file, "w", encoding="utf-8") as f:
                    json.dump({}, f)
                d = json.load(open(json_file, "r", encoding="utf-8"))
                data = {**d, str(n): hash_p}
                json.dump(data, open(json_file, "w", encoding="utf-8"), ensure_ascii=False, indent=4)
                print(hash_p)
        else:
            mb.showwarning('无法注册', '重复的密码不相同')
        with open('info/user.json', 'r', encoding='utf-8') as a:
            a = json.load(a)
    s = tk.Toplevel()
    s.geometry('400x400')
    s.title('注册')

    tk.Label(s, text='Email:').place(x=100, y=70)
    e = tk.Entry(s, width=15)
    e.place(x=150, y=70)
    tk.Label(s, text='用户名:').place(x=100, y=120)
    n = tk.Entry(s, width=15)
    n.place(x=150, y=120)
    tk.Label(s, text='密码:').place(x=100, y=170)
    p = tk.Entry(s, width=15)
    p.place(x=150, y=170)
    tk.Label(s, text='重复密码:').place(x=90, y=220)
    p2 = tk.Entry(s, width=15)
    p2.place(x=150, y=220)
    ttk.Button(s, text='OK', command=create_account).place(x=150, y=280)

    s.mainloop()
def ok():
    global l
    try:
        with open('info/user.json', 'r', encoding='utf-8') as u:
            ud = json.load(u)
    except:
        ud = {}

    input_name = name.get()
    input_pwd_raw = pw.get()
    input_pwd_sha256 = hashlib.sha256(input_pwd_raw.encode('utf-8')).hexdigest()

    if not ud:
        mb.showwarning('登录失败', '没注册账号')
        return

    login_success = False
    for username, user_data in ud.items():
        if (input_name == username or input_name == user_data.get('email')) and user_data.get(
                'password') == input_pwd_sha256:
            login_success = True
            break

    if login_success:
        main_sg()
        if l.winfo_exists():
            l.destroy()
    else:
        mb.showwarning('登录失败', '用户名/邮箱/密码错误')
def main_sg():
    sg=tk.Tk()
    sg.title('学生成绩管理系统')
    sg.geometry('400x300')

    def xuesheng():
        sg.geometry('750x400')
        def save(omg=114514):
            with open('info/student.json','w',encoding='utf-8') as f:
                rs=s.get_children()
                d={}
                for i in rs:
                    items=list(s.item(i,'values'))
                    d[items[0]]=items[1:]
                json.dump(d,f,ensure_ascii=False, indent=4)
        def read(omg=114514):
            if s.selection():
                sed=s.item(s.selection()[0],'values')
                ID.delete(0, 'end')
                name.delete(0, 'end')
                class_.delete(0, 'end')
                gander.delete(0, 'end')
                address.delete(0, 'end')
                phone.delete(0, 'end')
                for idx, val in enumerate(sed):
                    if idx == 0:
                        ID.insert(0, val)
                    elif idx == 1:
                        name.insert(0, val)
                    elif idx == 2:
                        class_.set(val)
                    elif idx == 3:
                        gander.set(val)
                    elif idx == 4:
                        address.insert(0, val)
                    elif idx == 5:
                        phone.insert(0, val)
        def add(omg=114514):
            idd=ID.get()
            if not idd:
                idd=str(len(s.get_children())+1)
            if name.get().strip() and class_.get().strip():
                s.insert('', 'end', values=(idd,name.get(),class_.get(),gander.get(),address.get(),phone.get()),iid=idd)
                save()
            else:
                mb.showwarning('无法添加','姓名/班级未填')
        def edit():
            idd = ID.get()
            if name.get().strip() and class_.get().strip() and s.selection():
                s.item(s.selection()[0], values=(idd, name.get(), class_.get(), gander.get(), address.get(), phone.get()))
                save()
            else:
                mb.showwarning('无法编辑', '姓名或班级未填/未选中行')
        def delete(omg=114514):
            if s.selection():
                for i in s.selection():
                    s.delete(i)
            else:
                mb.showwarning('无法删除','未选中行')
        def export():
            fp=fd.asksaveasfilename(title='导出',defaultextension='.xlsx',filetypes=[("Excel文档", ".xlsx"),("学生信息文件", ".stu")],initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            if fp:
                file_name, file_ext = os.path.splitext(fp)
                if file_ext.lower() == '.stu':
                    shutil.copy(r'info/student.json', fp)
                elif file_ext.lower() == '.xlsx':
                    wb=openpyxl.Workbook()
                    sheet=wb.active
                    sheet.title = '学生信息'

                    sheet['A1']='学生姓名'
                    sheet['B1']='班级'
                    sheet['C1']='性别'
                    sheet['D1']='家庭住址'
                    sheet['E1']='联系电话'
                    sheet.column_dimensions['A'].width = 10.68
                    sheet.column_dimensions['B'].width = 10.89
                    sheet.column_dimensions['D'].width = 26.7
                    sheet.column_dimensions['E'].width = 19.76
                    rows=[]
                    for i in s.get_children():
                        r=s.item(i)['values'][1:]
                        rows.append(r)
                    for row in rows:
                        sheet.append(row)

                    wb.save(fp)
        def import_():
            fp=fd.askopenfilename(title='导入',filetypes=[("学生信息文件", ".stu"),("Excel文档", ".xlsx")],initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            file_name, file_ext = os.path.splitext(fp)
            if file_ext.lower() == '.stu':
                try:
                    with open(fp, 'r', encoding='utf-8') as ss:
                        stu = json.load(ss)
                    for idd, vs in stu.items():
                        s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
                except Exception as e:
                    mb.showwarning('无法打开',str(e))
            elif file_ext.lower() == '.xlsx':
                s.delete(*s.get_children())
                wb=openpyxl.load_workbook(fp)
                sheet=wb.worksheets[0]
                titles=sheet[1]
                s.delete(*s.get_children())
                name_col = None
                class_col = None
                gender_col = None
                address_col = None
                phone_col = None

                for i in titles:
                    if '姓名' in str(i.value) or '名字' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        name_col = list(sheet[col_letter])
                        name_col.pop(0)

                    if '班级' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        class_col = list(sheet[col_letter])
                        class_col.pop(0)

                    if '性别' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        gender_col = list(sheet[col_letter])
                        gender_col.pop(0)

                    if '住址' in str(i.value) or '地址' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        address_col = list(sheet[col_letter])
                        address_col.pop(0)

                    if '电话' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        phone_col = list(sheet[col_letter])
                        phone_col.pop(0)

                idd = 1
                max_len = max(
                    len(name_col or []),
                    len(class_col or []),
                    len(gender_col or []),
                    len(address_col or []),
                    len(phone_col or [])
                )

                for i in range(max_len):
                    name_val = name_col[i].value if (name_col and i < len(name_col)) else ""
                    class_val = class_col[i].value if (class_col and i < len(class_col)) else ""
                    gender_val = gender_col[i].value if (gender_col and i < len(gender_col)) else ""
                    address_val = address_col[i].value if (address_col and i < len(address_col)) else ""
                    phone_val = phone_col[i].value if (phone_col and i < len(phone_col)) else ""

                    s.insert('', 'end', values=(idd, name_val, class_val, gender_val, address_val, phone_val), iid=idd)
                    idd += 1
        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)
        quit_()

        bj = []
        try:
            with open('info/class.json', 'r', encoding='utf-8') as c:
                cls = json.load(c)
            for i in cls.values():
                bj.append(''.join(reversed(i)))
        except:
            with open('info/class.json', 'w', encoding='utf-8') as c:
                c.write('{}')

        s = ttk.Treeview(sg, columns=('ID', 'name', 'class', 'gander', 'address','phone'),show='headings')
        scroll_c = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=s.yview)
        scroll_c.pack(side=tk.RIGHT, fill=tk.Y)
        s.config(yscrollcommand=scroll_c.set)
        s.pack(anchor='nw')
        s.heading('ID', text='学生编号', anchor='center', )
        s.heading('name', text='学生姓名', anchor='center')
        s.heading('class', text='班级', anchor='center')
        s.heading('gander',text='性别',anchor='center')
        s.heading('address',text='家庭住址',anchor='center')
        s.heading('phone',text='联系电话',anchor='center')
        s.column('ID', width=80, anchor='center')
        s.column('name', width=80, anchor='center')
        s.column('class', width=80, anchor='center')
        s.column('gander', width=60, anchor='center')
        s.column('address', width=200, anchor='center')
        s.column('phone', width=100, anchor='center')

        try:
            with open('info/student.json', 'r', encoding='utf-8') as ss:
                stu=json.load(ss)
            for idd,vs in stu.items():
                s.insert('', 'end', values=(idd,vs[0],vs[1],vs[2],vs[3],vs[4]),iid=idd)
        except:
            with open('info/student.json', 'w', encoding='utf-8') as f:
                f.write('{}')

        tk.Label(sg, text="学生编号：").place(x=7, y=260)
        ID = tk.Entry(sg, width=10)
        ID.place(x=80, y=260)
        tk.Label(sg, text="学生姓名：").place(x=167, y=260)
        name = tk.Entry(sg, width=15)
        name.place(x=240, y=260)
        tk.Label(sg, text="班级：").place(x=367, y=260)
        class_ = ttk.Combobox(sg, width=9,values=bj)
        class_.place(x=405, y=260)
        tk.Label(sg, text="性别：").place(x=500, y=260)
        gander = ttk.Combobox(sg,values=['男','女'],width=5)
        gander.current(0)
        gander.place(x=550, y=260)
        tk.Label(sg, text="家庭住址：").place(x=7, y=300)
        address = tk.Entry(sg, width=40)
        address.place(x=80, y=300)
        tk.Label(sg, text="联系电话：").place(x=430, y=300)
        phone = tk.Entry(sg, width=15)
        phone.place(x=500, y=300)

        ttk.Button(sg, text='添加', command=add).place(x=630, y=5)
        ttk.Button(sg, text='修改', command=edit).place(x=630, y=57)
        ttk.Button(sg, text='删除', command=delete).place(x=630, y=110)
        ttk.Button(sg, text='导入', command=import_).place(x=630, y=158)
        ttk.Button(sg, text='导出', command=export).place(x=630, y=210)
        sg.bind('<<TreeviewSelect>>', read)
        sg.bind('<Return>', add)
        sg.bind('<Escape>', quit_)
        sg.bind('<Delete>', delete)
    def chengji():
        sg.geometry('750x400')

        def save(omg=114514):
            with open('info/score.json', 'w', encoding='utf-8') as f:
                rs = s.get_children()
                d = {}
                for i in rs:
                    items = list(s.item(i, 'values'))
                    d[items[0]] = items[1:]
                json.dump(d, f, ensure_ascii=False, indent=4)
        def read(omg=114514):
            if s.selection():
                sed = s.item(s.selection()[0], 'values')
                ID.delete(0, 'end')
                name.delete(0, 'end')
                class_.delete(0, 'end')
                gander.delete(0, 'end')
                address.delete(0, 'end')
                phone.delete(0, 'end')
                ID.insert(0, sed[0])
                name.insert(0, sed[1])
                class_.set(sed[2])
                gander.set(sed[3])
                address.insert(0, sed[4])
                phone.insert(0, sed[5])
        def add(omg=114514):
            idd = ID.get()
            if not idd:
                idd = str(len(s.get_children()) + 1)
            if name.get().strip() and class_.get().strip() and gander.get().strip() and address.get().strip() and phone.get().strip():
                s.insert('', 'end', values=(idd, name.get(), class_.get(), gander.get(), address.get(), phone.get()),
                         iid=idd)
                save()
            else:
                mb.showwarning('无法添加', '姓名/班级未填')
        def edit():
            idd = ID.get()
            if name.get().strip() and class_.get().strip() and s.selection():
                s.item(s.selection()[0],
                       values=(idd, name.get(), class_.get(), gander.get(), address.get(), phone.get()))
                save()
            else:
                mb.showwarning('无法编辑', '姓名或班级未填/未选中行')
        def delete(omg=114514):
            if s.selection():
                for i in s.selection():
                    s.delete(i)
                    save()
            else:
                mb.showwarning('无法删除', '未选中行')

        def export():
            fp = fd.asksaveasfilename(title='导出', defaultextension='.xlsx',
                                      filetypes=[("Excel文档", ".xlsx"), ("成绩信息文件", ".sg")],
                                      initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            if fp:
                file_name, file_ext = os.path.splitext(fp)
                if file_ext.lower() == '.sg':
                    shutil.copy(r'info/score.json', fp)
                elif file_ext.lower() == '.xlsx':
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = '成绩信息'

                    sheet['A1'] = '学生姓名'
                    sheet['B1'] = '班级'
                    sheet['C1'] = '科目'
                    sheet['D1'] = '考试种类'
                    sheet['E1'] = '成绩'
                    sheet.column_dimensions['A'].width = 10.68
                    sheet.column_dimensions['B'].width = 10.89
                    sheet.column_dimensions['D'].width = 19.76
                    rows = []
                    for i in s.get_children():
                        r = s.item(i)['values'][1:]
                        rows.append(r)
                    for row in rows:
                        sheet.append(row)

                    wb.save(fp)

        def import_():
            fp = fd.askopenfilename(title='导入', filetypes=[("成绩信息文件", ".sg"), ("Excel文档", ".xlsx")],
                                    initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            file_name, file_ext = os.path.splitext(fp)
            if file_ext.lower() == '.sg':
                try:
                    with open(fp, 'r', encoding='utf-8') as ss:
                        stu = json.load(ss)
                    for idd, vs in stu.items():
                        s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
                except Exception as e:
                    mb.showwarning('无法打开', str(e))
            elif file_ext.lower() == '.xlsx':
                s.delete(*s.get_children())
                wb = openpyxl.load_workbook(fp)
                sheet = wb.worksheets[0]
                titles = sheet[1]
                s.delete(*s.get_children())
                name_col = None
                class_col = None
                km_col = None
                kszl_col = None
                cj_col = None

                for i in titles:
                    if '姓名' in str(i.value) or '名字' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        name_col = list(sheet[col_letter])
                        name_col.pop(0)

                    if '班级' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        class_col = list(sheet[col_letter])
                        class_col.pop(0)

                    if '科目' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        km_col = list(sheet[col_letter])
                        km_col.pop(0)

                    if '考试种类' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        kszl_col = list(sheet[col_letter])
                        kszl_col.pop(0)

                    if '成绩' in str(i.value):
                        col_letter = utils.get_column_letter(titles.index(i) + 1)
                        cj_col = list(sheet[col_letter])
                        cj_col.pop(0)

                idd = 1
                max_len = max(
                    len(name_col or []),
                    len(class_col or []),
                    len(km_col or []),
                    len(kszl_col or []),
                    len(cj_col or [])
                )

                for i in range(max_len):
                    name_val = name_col[i].value if (name_col and i < len(name_col)) else ""
                    class_val = class_col[i].value if (class_col and i < len(class_col)) else ""
                    km_val = km_col[i].value if (km_col and i < len(km_col)) else ""
                    kszl_val = kszl_col[i].value if (kszl_col and i < len(kszl_col)) else ""
                    cj_val = cj_col[i].value if (cj_col and i < len(cj_col)) else ""

                    s.insert('', 'end', values=(idd, name_val, class_val, km_val, kszl_val, cj_val), iid=idd)
                    idd += 1
        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)
        quit_()

        bj = []
        xs=[]
        ksty=[]
        km=[]
        try:
            with open('info/class.json', 'r', encoding='utf-8') as c:
                cls = json.load(c)
            for i in cls.values():
                bj.append(''.join(reversed(i)))
        except:
            with open('info/class.json', 'w', encoding='utf-8') as c:
                c.write('{}')
        try:
            with open('info/student.json', 'r', encoding='utf-8') as c:
                cls = json.load(c)
            for i in cls.values():
                xs.append(i[0])
        except:
            with open('info/student.json', 'w', encoding='utf-8') as c:
                c.write('{}')
        try:
            with open('info/exam.json', 'r', encoding='utf-8') as c:
                cls = json.load(c)
            for i in cls.values():
                ksty.append(i)
        except:
            with open('info/exam.json', 'w', encoding='utf-8') as c:
                c.write('{}')
        try:
            with open('info/subject.json', 'r', encoding='utf-8') as c:
                cls = json.load(c)
            for i in cls.values():
                km.append(i)
        except:
            with open('info/subject.json', 'w', encoding='utf-8') as c:
                c.write('{}')

        s = ttk.Treeview(sg, columns=('ID', 'name', 'class', 'gander', 'address', 'phone'), show='headings')
        scroll_c = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=s.yview)
        scroll_c.pack(side=tk.RIGHT, fill=tk.Y)
        s.config(yscrollcommand=scroll_c.set)
        s.pack(anchor='nw')
        s.heading('ID', text='编号', anchor='center', )
        s.heading('name', text='姓名', anchor='center')
        s.heading('class', text='班级', anchor='center')
        s.heading('gander', text='科目', anchor='center')
        s.heading('address', text='考试种类', anchor='center')
        s.heading('phone', text='成绩', anchor='center')
        s.column('ID', width=80, anchor='center')
        s.column('name', width=80, anchor='center')
        s.column('class', width=80, anchor='center')
        s.column('gander', width=60, anchor='center')
        s.column('address', width=200, anchor='center')
        s.column('phone', width=100, anchor='center')

        try:
            with open('info/score.json', 'r', encoding='utf-8') as ss:
                stu = json.load(ss)
            for idd, vs in stu.items():
                s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
        except:
            with open('info/score.json', 'w', encoding='utf-8') as f:
                f.write('{}')

        tk.Label(sg, text="编号：").place(x=7, y=260)
        ID = tk.Entry(sg, width=9)
        ID.place(x=80, y=260)
        tk.Label(sg, text="学生姓名：").place(x=167, y=260)
        name = ttk.Combobox(sg, width=14,values=xs)
        name.place(x=240, y=260)
        tk.Label(sg, text="班级：").place(x=367, y=260)
        class_ = ttk.Combobox(sg, width=10, values=bj)
        class_.place(x=405, y=260)
        tk.Label(sg, text="科目：").place(x=390, y=300)
        gander = ttk.Combobox(sg,width=7,values=km)
        gander.place(x=430, y=300)
        tk.Label(sg, text="考试种类：").place(x=7, y=300)
        address = ttk.Combobox(sg, width=15,values=ksty)
        address.place(x=80, y=300)
        tk.Label(sg, text="成绩：").place(x=220, y=300)
        phone = tk.Entry(sg, width=15)
        phone.place(x=270, y=300)

        ttk.Button(sg, text='添加', command=add).place(x=630, y=5)
        ttk.Button(sg, text='修改', command=edit).place(x=630, y=57)
        ttk.Button(sg, text='删除', command=delete).place(x=630, y=110)
        ttk.Button(sg, text='导入', command=import_).place(x=630, y=158)
        ttk.Button(sg, text='导出', command=export).place(x=630, y=210)
        sg.bind('<<TreeviewSelect>>', read)
        sg.bind('<Return>', add)
        sg.bind('<Escape>', quit_)
        sg.bind('<Delete>', delete)
    def nianji():
            def save(omg=114514):
                with open('info/grade.json', 'w', encoding='utf-8') as gd:
                    d = {}
                    if g.get_children():
                        for i in g.get_children():
                            d[g.item(i, 'values')[0]] = g.item(i, 'values')[1]
                    json.dump(d, gd)
            def read(omg=114514):
                if g.selection():
                    gid = g.item(g.selection()[0], 'values')[0]
                    gn = g.item(g.selection()[0], 'values')[1]
                    ID.delete(0, 'end')
                    ID.insert(0, gid)
                    n.delete(0, tk.END)
                    n.insert(0, gn)
            def add(omg=114514):
                gid = ID.get().strip()
                gn = n.get().strip()
                if gn:
                    if not gid:
                        gid = len(g.get_children()) + 1
                    try:
                        g.insert("", "end", values=(gid, gn), iid=gid)
                        save()
                    except:
                        mb.showwarning('无法创建', '该ID已存在')
                else:
                    mb.showwarning('无法创建', '年级名称没填')
            def edit():
                gid = ID.get().strip()
                gn = n.get().strip()
                if gn:
                    if not gid:
                        gid = len(g.get_children()) + 1
                    try:
                        g.item(g.selection()[0], values=(gid, gn))
                        save()
                    except:
                        mb.showwarning('无法创建', '该ID已存在')
                else:
                    mb.showwarning('无法创建', '年级名称没填')
            def delete(omg=114514):
                for i in g.selection():
                    g.delete(i)
                save()
            def export():
                p = fd.asksaveasfilename(title='导出', filetypes=[("年级信息文件", "*.gdf")], defaultextension='.gdf',
                                         initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
                try:
                    shutil.copy(r'info/grade.json', p)
                except Exception as e:
                    mb.showerror('无法保存', str(e))
            def import_():
                p = fd.askopenfilename(title='导入', filetypes=[("年级信息文件", "*.gdf")],
                                       initialdir=os.path.expanduser("~") + "/Desktop")
                for item in g.get_children(""):
                    g.delete(item)
                with open(p, 'r', encoding='utf-8') as f:
                    a = json.load(f)
                    for x, y in a.items():
                        g.insert("", "end", values=(x, y), iid=x)
            def quit_(omg=114514):
                def destroy_widgets(parent):
                    for widget in reversed(parent.winfo_children()):
                        if widget is tips or isinstance(widget, tk.Menu):
                            continue
                        destroy_widgets(widget)
                        try:
                            widget.destroy()
                        except:
                            pass

                destroy_widgets(sg)

            quit_()
            g = ttk.Treeview(sg, columns=('ID', 'name',), show='headings')
            scroll_g = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=g.yview)
            scroll_g.pack(side=tk.RIGHT, fill=tk.Y)
            g.config(yscrollcommand=scroll_g.set)
            g.pack(anchor='nw')
            g.heading('ID', text='年级编号', anchor='center')
            g.heading('name', text='年级名称', anchor='center')
            ttk.Button(sg, text='添加', command=add).place(x=20, y=260)
            ttk.Button(sg, text='修改', command=edit).place(x=112, y=260)
            ttk.Button(sg, text='删除', command=delete).place(x=205, y=260)
            ttk.Button(sg, text='导入', command=import_).place(x=300, y=260)
            ttk.Button(sg, text='导出', command=export).place(x=395, y=260)
            try:
                with open('info/grade.json', 'r', encoding='utf-8') as f:
                    a = json.load(f)
                    for x, y in a.items():
                        g.insert("", "end", values=(x, y), iid=x)
            except:
                with open('info/grade.json', 'w', encoding='utf-8') as f:
                    f.write('{}')
            sg.geometry('520x350')
            ttk.Label(sg, text='年级编号:\n\n\n\n\n年级名称:').place(x=430, y=20)
            ID = tk.Entry(sg, width=10)
            ID.place(x=430, y=40)
            n = tk.Entry(sg, width=10)
            n.place(x=430, y=130)
            sg.bind('<<TreeviewSelect>>', read)
            sg.bind('<Return>', add)
            sg.bind('<Escape>', quit_)
            sg.bind('<Delete>', delete)
    def banji():
        def save(omg=114514):
            with open('info/class.json', 'w', encoding='utf-8') as gd:
                if c.get_children():
                    d = {}
                    for i in c.get_children():
                        d[c.item(i, 'values')[0]] = [c.item(i, 'values')[1], c.item(i, 'values')[2]]
                json.dump(d, gd)
        def read(omg=114514):
            if c.selection():
                cid = c.item(c.selection()[0], 'values')[0]
                cn = c.item(c.selection()[0], 'values')[1]
                cg = c.item(c.selection()[0], 'values')[2]
                ID.delete(0, 'end')
                ID.insert(0, cid)
                n.delete(0, tk.END)
                n.insert(0, cn)
                cl.set(cg)
        def add(omg=114514):
            cid = ID.get().strip()
            cn = n.get().strip()
            cg = cl.get().strip()
            if cn and cg:
                if not cid:
                    cid = len(c.get_children()) + 1
                try:
                    c.insert("", "end", values=(cid, cn, cg), iid=cid)
                    save()
                except:
                    mb.showwarning('无法创建', '该ID已存在')
            else:
                mb.showwarning('无法创建', '班级名称/所属年级没填')
        def edit():
            cid = ID.get().strip()
            cn = n.get().strip()
            cg = cl.get().strip()
            if cn and cg:
                if not cid:
                    cid = len(c.get_children()) + 1
                c.item(c.selection()[0], values=(cid, cn, cg))
                save()
            else:
                mb.showwarning('无法创建', '班级名称/所属年级没填')
        def delete(omg=114514):
            for i in c.selection():
                c.delete(i)
            save()
        def export():
            p = fd.asksaveasfilename(title='导出', filetypes=[("班级信息文件", "*.cdf")], defaultextension='.cdf',
                                     initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            try:
                shutil.copy(r'info/class.json', p)
            except Exception as e:
                mb.showerror('无法保存', str(e))
        def import_():
            p = fd.askopenfilename(title='导入', filetypes=[("班级信息文件", "*.cdf")],
                                   initialdir=os.path.expanduser("~") + "/Desktop")
            for item in c.get_children(""):
                c.delete(item)
            with open(p, 'r', encoding='utf-8') as f:
                a = json.load(f)
                for x, y in a.items():
                    c.insert("", "end", values=(x, y[1], y[2]), iid=x)
        def quit_(omg=114514):
            for widget in sg.winfo_children():
                if isinstance(widget, tk.Menu):
                    continue
                if widget is tips:
                    continue
                widget.destroy()

        quit_()
        c = ttk.Treeview(sg, columns=('ID', 'grade', 'class'), show='headings')
        scroll_c = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=c.yview)
        scroll_c.pack(side=tk.RIGHT, fill=tk.Y)
        c.config(yscrollcommand=scroll_c.set)
        c.pack(anchor='nw')
        c.heading('ID', text='班级编号', anchor='center', )
        c.heading('grade', text='年级', anchor='center')
        c.heading('class', text='班级', anchor='center')
        c.column('class', width=130)
        c.column('grade', width=130)
        c.column('ID', width=130)
        ttk.Button(sg, text='添加', command=add).place(x=20, y=260)
        ttk.Button(sg, text='修改', command=edit).place(x=112, y=260)
        ttk.Button(sg, text='删除', command=delete).place(x=205, y=260)
        ttk.Button(sg, text='导入', command=import_).place(x=300, y=260)
        ttk.Button(sg, text='导出', command=export).place(x=395, y=260)
        try:
            with open('info/class.json', 'r', encoding='utf-8') as f:
                a = json.load(f)
                for i, j in a.items():
                    c.insert("", "end", values=(i, j[1], j[0]), iid=i)
        except:
            with open('info/class.json', 'w', encoding='utf-8') as f:
                f.write('{}')
        nianji_list = []
        try:
            with open('info/grade.json', 'r', encoding='utf-8') as f:
                a = json.load(f)
                if a:
                    for gs in a.values():
                        nianji_list.append(gs)
                else:
                    nianji()
                    mb.showinfo('年级未创建', '年级未创建，请先创建')
        except:
            with open('info/grade.json', 'w', encoding='utf-8') as f:
                f.write('{}')
                nianji()
                mb.showinfo('年级未创建', '年级未创建，请先创建')
        sg.geometry('520x350')
        ttk.Label(sg, text='班级编号:\n\n\n\n班级名称:\n\n\n\n所属年级:').place(x=430, y=20)
        ID = tk.Entry(sg, width=10)
        ID.place(x=430, y=40)
        n = tk.Entry(sg, width=10)
        n.place(x=430, y=110)
        cl = ttk.Combobox(sg, width=8, values=nianji_list)
        cl.place(x=430, y=180)
        sg.bind('<<TreeviewSelect>>', read)
        sg.bind('<Return>', add)
        sg.bind('<Escape>', quit_)
        sg.bind('<Delete>', delete)
    def kemu():
        def save(omg=114514):
            with open('info/subject.json', 'w', encoding='utf-8') as gd:
                d = {}
                if s.get_children():
                    for i in s.get_children():
                        d[s.item(i, 'values')[0]] = s.item(i, 'values')[1]
                json.dump(d, gd)
        def read(omg=114514):
            if s.selection():
                gid = s.item(s.selection()[0], 'values')[0]
                gn = s.item(s.selection()[0], 'values')[1]
                ID.delete(0, 'end')
                ID.insert(0, gid)
                n.delete(0, tk.END)
                n.insert(0, gn)
        def add(omg=114514):
            gid = ID.get().strip()
            gn = n.get().strip()
            if gn:
                if not gid:
                    gid = len(s.get_children()) + 1
                try:
                    s.insert("", "end", values=(gid, gn), iid=gid)
                    save()
                except:
                    mb.showwarning('无法创建', '该ID已存在')
            else:
                mb.showwarning('无法创建', '科目名称没填')
        def edit():
            gid = ID.get().strip()
            gn = n.get().strip()
            if gn:
                if not gid:
                    gid = len(s.get_children()) + 1
                try:
                    s.item(s.selection()[0], values=(gid, gn))
                    save()
                except:
                    mb.showwarning('无法创建', '该ID已存在')
            else:
                mb.showwarning('无法创建', '科目名称没填')
        def delete(omg=114514):
            for i in s.selection():
                s.delete(i)
            save()
        def export():
            p = fd.asksaveasfilename(title='导出', filetypes=[("科目信息文件", "*.sdf")], defaultextension='.sdf',
                                     initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            try:
                shutil.copy(r'info/subject.json', p)
            except Exception as e:
                mb.showerror('无法保存', str(e))
        def import_():
            p = fd.askopenfilename(title='导入', filetypes=[("科目信息文件", "*.sdf")],
                                   initialdir=os.path.expanduser("~") + "/Desktop")
            for item in s.get_children(""):
                s.delete(item)
            with open(p, 'r', encoding='utf-8') as f:
                a = json.load(f)
                for x, y in a.items():
                    s.insert("", "end", values=(x, y), iid=x)

        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)

        quit_()
        s = ttk.Treeview(sg, columns=('ID', 'name',), show='headings')
        scroll_s = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=s.yview)
        scroll_s.pack(side=tk.RIGHT, fill=tk.Y)
        s.config(yscrollcommand=scroll_s.set)
        s.pack(anchor='nw')
        s.heading('ID', text='科目编号', anchor='center')
        s.heading('name', text='科目名称', anchor='center')
        ttk.Button(sg, text='添加', command=add).place(x=20, y=260)
        ttk.Button(sg, text='修改', command=edit).place(x=112, y=260)
        ttk.Button(sg, text='删除', command=delete).place(x=205, y=260)
        ttk.Button(sg, text='导入', command=import_).place(x=300, y=260)
        ttk.Button(sg, text='导出', command=export).place(x=395, y=260)
        try:
            with open('info/subject.json', 'r', encoding='utf-8') as f:
                a = json.load(f)
                for x, y in a.items():
                    s.insert("", "end", values=(x, y), iid=x)
        except:
            with open('info/subject.json', 'w', encoding='utf-8') as f:
                f.write('{}')
        sg.geometry('520x350')
        ttk.Label(sg, text='科目编号:\n\n\n\n\n科目名称:').place(x=430, y=20)
        ID = tk.Entry(sg, width=10)
        ID.place(x=430, y=40)
        n = tk.Entry(sg, width=10)
        n.place(x=430, y=130)
        sg.bind('<<TreeviewSelect>>', read)
        sg.bind('<Return>', add)
        sg.bind('<Escape>', quit_)
        sg.bind('<Delete>', delete)
    def kaoshizhonglei():
        def save(omg=114514):
            with open('info/exam.json', 'w', encoding='utf-8') as gd:
                d = {}
                if e.get_children():
                    for i in e.get_children():
                        d[e.item(i, 'values')[0]] = e.item(i, 'values')[1]
                json.dump(d, gd)
        def read(omg=114514):
            if e.selection():
                gid = e.item(e.selection()[0], 'values')[0]
                gn = e.item(e.selection()[0], 'values')[1]
                ID.delete(0, 'end')
                ID.insert(0, gid)
                n.delete(0, tk.END)
                n.insert(0, gn)
        def add(omg=114514):
            gid = ID.get().strip()
            gn = n.get().strip()
            if gn:
                if not gid:
                    gid = len(e.get_children()) + 1
                try:
                    e.insert("", "end", values=(gid, gn), iid=gid)
                    save()
                except:
                    mb.showwarning('无法创建', '该ID已存在')
            else:
                mb.showwarning('无法创建', '考试名称没填')
        def edit():
            gid = ID.get().strip()
            gn = n.get().strip()
            if gn:
                if not gid:
                    gid = len(e.get_children()) + 1
                try:
                    e.item(e.selection()[0], values=(gid, gn))
                    save()
                except:
                    mb.showwarning('无法创建', '该ID已存在')
            else:
                mb.showwarning('无法创建', '考试类型名称没填')
        def delete(omg=114514):
            for i in e.selection():
                e.delete(i)
            save()
        def export():
            p = fd.asksaveasfilename(title='导出', filetypes=[("考试类型信息文件", "*.etdf")], defaultextension='.etdf',
                                     initialdir=os.path.join(os.path.expanduser("~"), "Desktop"))
            try:
                shutil.copy(r'info/exam.json', p)
            except Exception as e:
                mb.showerror('无法保存', str(e))
        def import_():
            p = fd.askopenfilename(title='导入', filetypes=[("考试类型信息文件", "*.etdf")],
                                   initialdir=os.path.expanduser("~") + "/Desktop")
            for item in e.get_children(""):
                e.delete(item)
            with open(p, 'r', encoding='utf-8') as f:
                a = json.load(f)
                for x, y in a.items():
                    e.insert("", "end", values=(x, y), iid=x)
        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)

        quit_()
        e = ttk.Treeview(sg, columns=('ID', 'name'), show='headings')
        scroll_e = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=e.yview)
        scroll_e.pack(side=tk.RIGHT,fill=tk.Y)
        e.config(yscrollcommand=scroll_e.set)
        e.pack(anchor='nw')
        e.heading('ID', text='科目编号', anchor='center')
        e.heading('name', text='科目名称', anchor='center')
        ttk.Button(sg, text='添加', command=add).place(x=20, y=260)
        ttk.Button(sg, text='修改', command=edit).place(x=112, y=260)
        ttk.Button(sg, text='删除', command=delete).place(x=205, y=260)
        ttk.Button(sg, text='导入', command=import_).place(x=300, y=260)
        ttk.Button(sg, text='导出', command=export).place(x=395, y=260)
        try:
            with open('info/exam.json', 'r', encoding='utf-8') as f:
                a = json.load(f)
                for x, y in a.items():
                    e.insert("", "end", values=(x, y), iid=x)
        except:
            with open('info/exam.json', 'w', encoding='utf-8') as f:
                f.write('{}')
        sg.geometry('520x350')
        ttk.Label(sg, text='考试种类编号:\n\n\n\n\n考试种类名称:').place(x=430, y=20)
        ID = tk.Entry(sg, width=10)
        ID.place(x=430, y=40)
        n = tk.Entry(sg, width=10)
        n.place(x=430, y=130)
        sg.bind('<<TreeviewSelect>>', read)
        sg.bind('<Return>', add)
        sg.bind('<Escape>', quit_)
        sg.bind('<Delete>', delete)
    def xschaxun():
        sg.geometry('780x320')

        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)
        def cx(omg=114514):
            sh=ID.get().strip()
            od=[]
            for i in s.get_children():
                row = s.item(i)["values"]
                od.append(row)
            if sh:
                for item in s.get_children():
                    s.delete(item)
                for data in od:
                    if sh in str(data).lower():
                        s.insert("", "end", values=data)
                mb.showinfo('查询',f'共查询到{len(s.get_children())}个结果')
            else:
                try:
                    with open('info/student.json', 'r', encoding='utf-8') as ss:
                        stu = json.load(ss)
                    for idd, vs in stu.items():
                        s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
                except:
                    with open('info/student.json', 'w', encoding='utf-8') as f:
                        f.write('{}')
        quit_()

        s = ttk.Treeview(sg, columns=('ID', 'name', 'class', 'gander', 'address','phone'),show='headings')
        scroll_c = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=s.yview)
        scroll_c.pack(side=tk.RIGHT, fill=tk.Y)
        s.config(yscrollcommand=scroll_c.set)
        s.pack(anchor='nw')
        s.heading('ID', text='学生编号', anchor='center', )
        s.heading('name', text='学生姓名', anchor='center')
        s.heading('class', text='班级', anchor='center')
        s.heading('gander',text='性别',anchor='center')
        s.heading('address',text='家庭住址',anchor='center')
        s.heading('phone',text='联系电话',anchor='center')
        s.column('ID', width=80, anchor='center')
        s.column('name', width=80, anchor='center')
        s.column('class', width=80, anchor='center')
        s.column('gander', width=60, anchor='center')
        s.column('address', width=200, anchor='center')
        s.column('phone', width=100, anchor='center')

        try:
            with open('info/student.json', 'r', encoding='utf-8') as ss:
                stu=json.load(ss)
            for idd,vs in stu.items():
                s.insert('', 'end', values=(idd,vs[0],vs[1],vs[2],vs[3],vs[4]),iid=idd)
        except:
            with open('info/student.json', 'w', encoding='utf-8') as f:
                f.write('{}')

        tk.Label(sg, text="关键字：").place(x=610, y=20)
        ID = tk.Entry(sg, width=10)
        ID.place(x=670, y=20)
        tk.Label(sg, text="范围：").place(x=610, y=70)
        class_ = ttk.Combobox(sg, width=9,values=['全部','学生编号', '学生姓名', '班级', '性别', '家庭住址', '联系电话'])
        class_.place(x=660, y=70)
        class_.current(0)
        ttk.Button(sg, text='查询', command=cx,padding=7).place(x=640, y=130)

        sg.bind('<Return>', cx)
        sg.bind('<Escape>', quit_)

    def cjchaxun():
        sg.geometry('780x320')

        def quit_(omg=114514):
            def destroy_widgets(parent):
                for widget in reversed(parent.winfo_children()):
                    if widget is tips or isinstance(widget, tk.Menu):
                        continue
                    destroy_widgets(widget)
                    try:
                        widget.destroy()
                    except:
                        pass

            destroy_widgets(sg)

        quit_()
        def cx(omg=114514):
            sh = ID.get().strip()
            od = []
            for i in s.get_children():
                row = s.item(i)["values"]
                od.append(row)
            fwei = class_.get()
            s.delete(*s.get_children())

            if fwei == "全部":
                for data in od:
                    if sh in str(data).lower():
                        s.insert("", "end", values=data)
            elif fwei == "编号":
                for data in od:
                    if sh in str(data[0]).lower():
                        s.insert("", "end", values=data)
            elif fwei == "学生姓名":
                for data in od:
                    if sh in str(data[1]).lower():
                        s.insert("", "end", values=data)
            elif fwei == "班级":
                for data in od:
                    if sh in str(data[2]).lower():
                        s.insert("", "end", values=data)
            elif fwei == "科目":
                for data in od:
                    if sh in str(data[3]).lower():
                        s.insert("", "end", values=data)
            elif fwei == "考试种类":
                for data in od:
                    if sh in str(data[4]).lower():
                        s.insert("", "end", values=data)
            elif fwei == "成绩":
                for data in od:
                    if sh in str(data[5]).lower():
                        s.insert("", "end", values=data)
            elif fwei == '分数段':
                def okk():
                    nonlocal sta,sto
                    sta=start.get()
                    sto=stop.get()
                    try:
                        min_score = float(sta)
                        max_score = float(sto)
                    except:
                        return

                    for data in od:
                        try:
                            score = float(data[5])
                            if min_score <= score <= max_score:
                                s.insert("", "end", values=data)
                        except:
                            pass
                    sas.destroy()
                sta,sto='',''
                sas = tk.Toplevel(sg)
                sas.title('分数段查询')
                sas.geometry('250x120')
                sas.resizable(0, 0)
                tk.Label(sas, text='请输入开始与结束分数').pack(pady=5)
                start = tk.Entry(sas)
                start.pack(pady=2)
                stop = tk.Entry(sas)
                stop.pack(pady=2)
                ttk.Button(sas, text='OK',padding=3,command=okk).pack(pady=5)
            else:
                try:
                    with open('info/score.json', 'r', encoding='utf-8') as ss:
                        stu = json.load(ss)
                    for idd, vs in stu.items():
                        s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
                except:
                    with open('info/score.json', 'w', encoding='utf-8') as f:
                        f.write('{}')

        s = ttk.Treeview(sg, columns=('ID', 'name', 'class', 'km', 'kszl', 'cj'), show='headings')
        scroll_c = ttk.Scrollbar(sg, orient=tk.VERTICAL, command=s.yview)
        scroll_c.pack(side=tk.RIGHT, fill=tk.Y)
        s.config(yscrollcommand=scroll_c.set)
        s.pack(anchor='nw')
        s.heading('ID', text='编号', anchor='center', )
        s.heading('name', text='学生姓名', anchor='center')
        s.heading('class', text='班级', anchor='center')
        s.heading('km', text='科目', anchor='center')
        s.heading('kszl', text='考试种类', anchor='center')
        s.heading('cj', text='成绩', anchor='center')
        s.column('ID', width=80, anchor='center')
        s.column('name', width=80, anchor='center')
        s.column('class', width=80, anchor='center')
        s.column('km', width=60, anchor='center')
        s.column('kszl', width=200, anchor='center')
        s.column('cj', width=100, anchor='center')

        try:
            with open('info/score.json', 'r', encoding='utf-8') as ss:
                stu = json.load(ss)
            for idd, vs in stu.items():
                s.insert('', 'end', values=(idd, vs[0], vs[1], vs[2], vs[3], vs[4]), iid=idd)
        except:
            pass

        tk.Label(sg, text="关键字：").place(x=610, y=20)
        ID = tk.Entry(sg, width=10)
        ID.place(x=670, y=20)
        tk.Label(sg, text="范围：").place(x=610, y=70)
        class_ = ttk.Combobox(sg, width=9,
                              values=['全部', '编号', '学生姓名', '班级', '科目', '考试种类', '成绩', '分数段'])
        class_.place(x=660, y=70)
        class_.current(0)
        ttk.Button(sg, text='查询', command=cx, padding=7).place(x=640, y=130)

        sg.bind('<Return>', cx)
        sg.bind('<Escape>', quit_)

    def fankui():
        web.open('hby-bge.mysxl.cn')

    tips=tk.Label(sg,text=f'当前用户：{name.get()} | 登录时间：{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())} | 作者:HBY')
    tips.pack(side='bottom',pady=5)
    menu = tk.Menu(sg)
    sets = tk.Menu(menu, tearoff=0)
    cx=tk.Menu(menu,tearoff=0)
    menu.add_cascade(label='基本数据管理',menu=sets)
    menu.add_command(label='学生管理',command=xuesheng)
    menu.add_command(label='成绩管理',command=chengji)
    menu.add_cascade(label='查询',menu=cx)
    cx.add_command(label='学生查询',command=xschaxun)
    cx.add_command(label='成绩查询', command=cjchaxun)
    sets.add_command(label='年级管理',command=nianji)
    sets.add_command(label='班级管理',command=banji)
    sets.add_command(label='科目管理',command=kemu)
    sets.add_command(label='考试种类管理',command=kaoshizhonglei)
    menu.add_command(label='反馈',command=fankui)

    sg.configure(menu=menu)
#TODO 主程序
l=tk.Tk()
l.geometry('400x400')
l.title('登录')

with open('info/user.json', 'r', encoding='utf-8') as u:
    a = json.load(u)

tk.Label(l,text='用户名/Email:').place(x=62,y=120)
name=tk.Entry(l,width=15)
name.place(x=150,y=120)
tk.Label(l,text='密码:').place(x=100,y=170)
pw=tk.Entry(l,width=15,show='*')
pw.place(x=150,y=170)
ttk.Button(l,text='没有账号？注册',command=sign_up).place(x=150,y=230)
ttk.Button(l,text='登录',command=ok).place(x=150,y=280)

l.mainloop()